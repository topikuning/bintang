"""Excel import service.

For each module, accepts an XLSX file, validates row-by-row, and either
returns a preview (dry-run) or commits to DB.

Multi-row entities (invoices, purchase orders) are grouped by a key column
so each unique value becomes one record with its line items.
"""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any, Awaitable, Callable

from openpyxl import Workbook, load_workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.models import (
    Category,
    CategoryType,
    Company,
    Invoice,
    InvoiceAttachment,
    InvoiceItem,
    InvoiceStatus,
    InvoiceType,
    PartyType,
    PaymentMethod,
    POItem,
    POStatus,
    Project,
    ProjectAttachment,
    ProjectStatus,
    PurchaseOrder,
    Transaction,
    TransactionAttachment,
    TxnStatus,
    TxnType,
    User,
    VendorClient,
    VendorClientType,
)
from app.services.storage.links import normalize_external_link


# ---------- Generic helpers ----------
def _str(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _parse_date(v: Any) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Tanggal tidak dikenali: {s}")


def _parse_decimal(v: Any) -> Decimal:
    if v is None or v == "":
        return Decimal("0")
    if isinstance(v, (int, float, Decimal)):
        return Decimal(str(v))
    s = str(v).strip().replace(" ", "").replace("Rp", "").replace("rp", "")
    # Indonesian: "1.000.000,50" → "1000000.50"; English: "1,000,000.50"
    if "," in s and "." in s:
        # if comma comes after dot, comma is decimal separator
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        # only comma -> assume Indonesian decimal
        if s.count(",") == 1 and len(s.split(",")[1]) <= 2:
            s = s.replace(",", ".")
        else:
            s = s.replace(",", "")
    try:
        return Decimal(s)
    except InvalidOperation:
        raise ValueError(f"Nilai numerik tidak valid: {v}")


def _parse_enum(value: Any, enum_cls, field_name: str):
    if value is None:
        raise ValueError(f"{field_name} wajib diisi")
    s = str(value).strip().upper()
    aliases = {
        "MASUK": "IN", "IN": "IN", "INCOME": "IN",
        "KELUAR": "OUT", "OUT": "OUT", "EXPENSE": "OUT",
        "VENDOR": "VENDOR", "CLIENT": "CLIENT", "BOTH": "BOTH",
        "AKTIF": "AKTIF", "ACTIVE": "AKTIF",
        "SELESAI": "SELESAI", "DONE": "SELESAI",
        "DITAHAN": "DITAHAN", "ON_HOLD": "DITAHAN", "HOLD": "DITAHAN",
        "DIBATALKAN": "DIBATALKAN", "CANCELLED": "DIBATALKAN",
        "CASH": "CASH", "TUNAI": "CASH",
        "TRANSFER": "TRANSFER", "TF": "TRANSFER",
        "QRIS": "QRIS", "GIRO": "GIRO", "OTHER": "OTHER", "LAINNYA": "OTHER",
        "COMPANY": "COMPANY", "PERUSAHAAN": "COMPANY",
        "PERSONAL": "PERSONAL", "EMPLOYEE": "EMPLOYEE", "KARYAWAN": "EMPLOYEE",
        "INTERNAL": "INTERNAL", "OPERASIONAL": "INTERNAL",
    }
    s = aliases.get(s, s)
    try:
        return enum_cls(s)
    except ValueError:
        valid = [e.value for e in enum_cls]
        raise ValueError(f"{field_name} '{value}' tidak valid (pilihan: {', '.join(valid)})")


def read_xlsx(content: bytes) -> list[dict[str, Any]]:
    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb.active
    if ws is None:
        return []
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
    out: list[dict[str, Any]] = []
    for r in rows[1:]:
        if all(c is None or str(c).strip() == "" for c in r):
            continue
        out.append({headers[i]: r[i] for i in range(len(headers))})
    return out


def build_template(headers: list[str], example: list[Any] | None = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    if ws is not None:
        ws.append(headers)
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)
        if example:
            ws.append(example)
        for i, h in enumerate(headers, 1):
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(i)].width = max(14, min(40, len(str(h)) + 4))
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------- Entity-specific importers ----------
async def _lookup(db: AsyncSession, model, **filters):
    stmt = select(model).filter_by(**filters)
    return (await db.execute(stmt)).scalar_one_or_none()


async def _lookup_ci(db: AsyncSession, model, field: str, value: str):
    """Case-insensitive lookup by a single string column."""
    col = getattr(model, field)
    stmt = select(model).where(func.lower(col) == value.lower())
    return (await db.execute(stmt)).scalar_one_or_none()


def _set_if_present(obj, attrs: dict) -> None:
    """Hanya tulis field yang TIDAK None (kolom kosong di Excel = jangan ditimpa)."""
    for k, v in attrs.items():
        if v is not None:
            setattr(obj, k, v)


# Action saat menemukan duplikat (dipakai oleh tiap handler):
#   "skip"   -> abaikan baris, klasifikasikan sebagai dupe
#   "update" -> overwrite field-data (status/audit tidak ikut)
#   "error"  -> raise; baris masuk errors
DupAction = str


async def import_companies(rows, db, user, *, commit, dup_action):
    """Natural key: lower(name)."""
    new, dupes, errors = [], [], []
    for i, r in enumerate(rows, start=2):
        try:
            name = _str(r.get("name"))
            if not name:
                raise ValueError("name wajib diisi")
            data = dict(
                address=_str(r.get("address")),
                npwp=_str(r.get("npwp")),
                phone=_str(r.get("phone")),
                email=_str(r.get("email")),
                director_name=_str(r.get("director_name")),
                bank_account=_str(r.get("bank_account")),
            )
            existing = await _lookup_ci(db, Company, "name", name)
            if existing and existing.deleted_at is None:
                if dup_action == "error":
                    raise ValueError(f"Perusahaan '{name}' sudah ada")
                if dup_action == "update" and commit:
                    _set_if_present(existing, data)
                dupes.append({"name": name, "action": dup_action})
                continue
            if commit:
                db.add(Company(name=name, **data))
            new.append({"name": name})
        except Exception as e:
            errors.append({"row": i, "message": str(e), "raw": r})
    return new, dupes, errors


async def import_categories(rows, db, user, *, commit, dup_action):
    """Natural key: (lower(name), type)."""
    new, dupes, errors = [], [], []
    for i, r in enumerate(rows, start=2):
        try:
            name = _str(r.get("name"))
            if not name:
                raise ValueError("name wajib diisi")
            ctype = _parse_enum(r.get("type"), CategoryType, "type")
            description = _str(r.get("description"))

            stmt = select(Category).where(
                func.lower(Category.name) == name.lower(),
                Category.type == ctype,
            )
            existing = (await db.execute(stmt)).scalar_one_or_none()
            if existing and existing.deleted_at is None:
                if dup_action == "error":
                    raise ValueError(f"Kategori '{name}' ({ctype.value}) sudah ada")
                if dup_action == "update" and commit:
                    _set_if_present(existing, {"description": description})
                dupes.append({"name": name, "type": ctype.value, "action": dup_action})
                continue
            if commit:
                db.add(Category(name=name, type=ctype, description=description))
            new.append({"name": name, "type": ctype.value})
        except Exception as e:
            errors.append({"row": i, "message": str(e), "raw": r})
    return new, dupes, errors


async def import_vendors_clients(rows, db, user, *, commit, dup_action):
    """Natural key: NPWP (kalau ada), fallback ke lower(name)."""
    new, dupes, errors = [], [], []
    for i, r in enumerate(rows, start=2):
        try:
            name = _str(r.get("name"))
            if not name:
                raise ValueError("name wajib diisi")
            vtype = VendorClientType.VENDOR
            if r.get("type"):
                vtype = _parse_enum(r.get("type"), VendorClientType, "type")
            data = dict(
                type=vtype,
                contact=_str(r.get("contact")),
                phone=_str(r.get("phone")),
                email=_str(r.get("email")),
                npwp=_str(r.get("npwp")),
                address=_str(r.get("address")),
                bank_account=_str(r.get("bank_account")),
            )
            existing = None
            if data["npwp"]:
                existing = await _lookup(db, VendorClient, npwp=data["npwp"])
            if not existing:
                existing = await _lookup_ci(db, VendorClient, "name", name)
            if existing and existing.deleted_at is None:
                if dup_action == "error":
                    raise ValueError(f"Vendor/Client '{name}' sudah ada")
                if dup_action == "update" and commit:
                    _set_if_present(existing, {"name": name, **data})
                dupes.append({"name": name, "action": dup_action})
                continue
            if commit:
                db.add(VendorClient(name=name, **data))
            new.append({"name": name, "type": vtype.value})
        except Exception as e:
            errors.append({"row": i, "message": str(e), "raw": r})
    return new, dupes, errors


async def import_projects(rows, db, user, *, commit, dup_action):
    """Natural key: code. Update ⇒ overwrite data field, status workflow tetap."""
    new, dupes, errors = [], [], []
    for i, r in enumerate(rows, start=2):
        try:
            code = _str(r.get("code"))
            name = _str(r.get("name"))
            company_name = _str(r.get("company_name"))
            if not (code and name and company_name):
                raise ValueError("code, name, company_name wajib")
            company = await _lookup_ci(db, Company, "name", company_name)
            if not company or company.deleted_at is not None:
                raise ValueError(f"Perusahaan '{company_name}' tidak ditemukan")
            pic_id = None
            if pic_email := _str(r.get("pic_email")):
                pic = await _lookup(db, User, email=pic_email)
                if not pic:
                    raise ValueError(f"User PIC '{pic_email}' tidak ditemukan")
                pic_id = pic.id
            status = None
            if r.get("status"):
                status = _parse_enum(r.get("status"), ProjectStatus, "status")
            project_value = _parse_decimal(r.get("project_value") or 0)
            budget_raw = r.get("budget_amount")
            if budget_raw in (None, ""):
                budget_amount = (project_value * Decimal("0.7")).quantize(Decimal("0.01"))
            else:
                budget_amount = _parse_decimal(budget_raw)
            doc_url = _str(r.get("document_url"))
            doc_label = _str(r.get("document_label"))
            doc_meta = None
            if doc_url:
                try:
                    doc_meta = normalize_external_link(doc_url, label=doc_label)
                except Exception as e:
                    raise ValueError(f"document_url tidak valid: {e}")
            data = dict(
                name=name,
                location=_str(r.get("location")),
                company_id=company.id,
                pic_user_id=pic_id,
                start_date=_parse_date(r.get("start_date")),
                end_date=_parse_date(r.get("end_date")),
                notes=_str(r.get("notes")),
                project_value=project_value,
                budget_amount=budget_amount,
                currency=_str(r.get("currency")) or "IDR",
                overbudget_tolerance_pct=_parse_decimal(r.get("overbudget_tolerance_pct") or 0),
            )

            existing = await _lookup(db, Project, code=code)
            if existing and existing.deleted_at is None:
                if dup_action == "error":
                    raise ValueError(f"Kode proyek '{code}' sudah ada")
                if dup_action == "update" and commit:
                    _set_if_present(existing, data)
                    if status is not None:
                        existing.status = status
                    if doc_meta:
                        db.add(ProjectAttachment(
                            project_id=existing.id, label=doc_label,
                            uploaded_by_id=user.id, **doc_meta,
                        ))
                dupes.append({"code": code, "name": name, "action": dup_action})
                continue
            if commit:
                p = Project(
                    code=code, status=status or ProjectStatus.AKTIF, **data,
                )
                db.add(p)
                if doc_meta:
                    await db.flush()
                    db.add(ProjectAttachment(
                        project_id=p.id, label=doc_label,
                        uploaded_by_id=user.id, **doc_meta,
                    ))
            new.append({"code": code, "name": name})
        except Exception as e:
            errors.append({"row": i, "message": str(e), "raw": r})
    return new, dupes, errors


async def import_transactions(rows, db, user, *, commit, dup_action):
    """Append-only. Dup detection: (project_id, tx_date, amount, reference_no).
    Default 'skip' agar tidak double-insert kalau user re-upload file yang sama.
    'update' tidak didukung -- lebih aman insert ulang setelah delete manual."""
    new, dupes, errors = [], [], []
    for i, r in enumerate(rows, start=2):
        try:
            project_code = _str(r.get("project_code"))
            if not project_code:
                raise ValueError("project_code wajib")
            project = await _lookup(db, Project, code=project_code)
            if not project or project.deleted_at is not None:
                raise ValueError(f"Proyek '{project_code}' tidak ditemukan")
            tx_date = _parse_date(r.get("tx_date"))
            if not tx_date:
                raise ValueError("tx_date wajib")
            ttype = _parse_enum(r.get("type"), TxnType, "type")
            amount = _parse_decimal(r.get("amount"))
            if amount <= 0:
                raise ValueError("amount harus > 0")
            cat_id = None
            if cn := _str(r.get("category_name")):
                cat = await _lookup_ci(db, Category, "name", cn)
                if not cat:
                    raise ValueError(f"Kategori '{cn}' tidak ditemukan")
                cat_id = cat.id
            vc_id = None
            if vn := _str(r.get("vendor_client_name")):
                vc = await _lookup_ci(db, VendorClient, "name", vn)
                if vc:
                    vc_id = vc.id
            method = PaymentMethod.TRANSFER
            if r.get("payment_method"):
                method = _parse_enum(r.get("payment_method"), PaymentMethod, "payment_method")
            party_type = None
            if r.get("party_type"):
                party_type = _parse_enum(r.get("party_type"), PartyType, "party_type")
            reference_no = _str(r.get("reference_no"))
            att_url = _str(r.get("attachment_url")) or _str(r.get("bukti_url"))
            att_label = _str(r.get("attachment_label")) or _str(r.get("bukti_label"))
            att_meta = None
            if att_url:
                try:
                    att_meta = normalize_external_link(att_url, label=att_label)
                except Exception as e:
                    raise ValueError(f"attachment_url tidak valid: {e}")

            # Dup detection -- hanya kalau reference_no diisi (kalau kosong terlalu
            # gampang false-positive)
            existing = None
            if reference_no:
                stmt = select(Transaction).where(
                    Transaction.project_id == project.id,
                    Transaction.tx_date == tx_date,
                    Transaction.amount == amount,
                    Transaction.reference_no == reference_no,
                    Transaction.deleted_at.is_(None),
                )
                existing = (await db.execute(stmt)).scalar_one_or_none()
            if existing:
                if dup_action == "error":
                    raise ValueError(
                        f"Transaksi {project_code} {tx_date} Rp{amount} ref={reference_no} sudah ada"
                    )
                # 'update' tidak didukung untuk transaksi -> diperlakukan sebagai 'skip'
                dupes.append({
                    "project": project_code, "tx_date": tx_date.isoformat(),
                    "amount": str(amount), "reference_no": reference_no,
                    "action": "skipped",
                })
                continue

            if commit:
                tx = Transaction(
                    project_id=project.id,
                    tx_date=tx_date, type=ttype,
                    category_id=cat_id, amount=amount,
                    party_type=party_type, party_name=_str(r.get("party_name")),
                    vendor_client_id=vc_id,
                    payment_method=method, reference_no=reference_no,
                    description=_str(r.get("description")),
                    status=TxnStatus.DRAFT, created_by_id=user.id,
                )
                db.add(tx)
                if att_meta:
                    await db.flush()
                    db.add(TransactionAttachment(
                        transaction_id=tx.id, uploaded_by_id=user.id, **att_meta,
                    ))
            new.append({
                "project": project_code, "tx_date": tx_date.isoformat(),
                "type": ttype.value, "amount": str(amount),
            })
        except Exception as e:
            errors.append({"row": i, "message": str(e), "raw": r})
    return new, dupes, errors


async def import_invoices(rows, db, user, *, commit, dup_action):
    """Multi-row per invoice, grouped by `number`.
    Natural key: (project_id, number). Update ⇒ replace items + overwrite header
    field, status & allocations TIDAK ikut diubah."""
    new, dupes, errors = [], [], []
    groups: dict[str, list[tuple[int, dict]]] = {}
    for i, r in enumerate(rows, start=2):
        num = _str(r.get("number"))
        if not num:
            errors.append({"row": i, "message": "number wajib", "raw": r})
            continue
        groups.setdefault(num, []).append((i, r))

    for num, group in groups.items():
        first_row_num, first = group[0]
        try:
            project_code = _str(first.get("project_code"))
            project = await _lookup(db, Project, code=project_code) if project_code else None
            if not project:
                raise ValueError(f"Proyek '{project_code}' tidak ditemukan")
            itype = _parse_enum(first.get("type"), InvoiceType, "type")
            inv_date = _parse_date(first.get("invoice_date"))
            if not inv_date:
                raise ValueError("invoice_date wajib")
            due = _parse_date(first.get("due_date"))
            tax = _parse_decimal(first.get("tax") or 0)
            vc_id = None
            if vn := _str(first.get("vendor_client_name")):
                vc = await _lookup_ci(db, VendorClient, "name", vn)
                if vc:
                    vc_id = vc.id

            items_data: list[dict] = []
            for row_no, r in group:
                desc = _str(r.get("item_description"))
                if not desc:
                    continue
                qty = _parse_decimal(r.get("item_quantity") or 1)
                price = _parse_decimal(r.get("item_unit_price") or 0)
                items_data.append(dict(
                    description=desc, quantity=qty,
                    unit=_str(r.get("item_unit")),
                    unit_price=price, subtotal=price * qty,
                ))
            subtotal = sum((it["subtotal"] for it in items_data), Decimal("0"))
            total = subtotal + tax
            att_url = _str(first.get("attachment_url")) or _str(first.get("bukti_url"))
            att_label = _str(first.get("attachment_label")) or _str(first.get("bukti_label"))
            att_meta = None
            if att_url:
                try:
                    att_meta = normalize_external_link(att_url, label=att_label)
                except Exception as e:
                    raise ValueError(f"attachment_url tidak valid: {e}")

            existing = (await db.execute(
                select(Invoice)
                .options(selectinload(Invoice.items))
                .where(
                    Invoice.project_id == project.id,
                    Invoice.number == num,
                    Invoice.deleted_at.is_(None),
                )
            )).scalar_one_or_none()
            if existing:
                if dup_action == "error":
                    raise ValueError(f"Invoice '{num}' di proyek {project_code} sudah ada")
                if dup_action == "update" and commit:
                    existing.invoice_date = inv_date
                    existing.due_date = due
                    existing.type = itype
                    existing.vendor_client_id = vc_id or existing.vendor_client_id
                    existing.party_name = _str(first.get("party_name")) or existing.party_name
                    existing.tax = tax
                    existing.subtotal = subtotal
                    existing.total = total
                    existing.notes = _str(first.get("notes")) or existing.notes
                    # replace items
                    existing.items.clear()
                    await db.flush()
                    for d in items_data:
                        existing.items.append(InvoiceItem(**d))
                    if att_meta:
                        db.add(InvoiceAttachment(
                            invoice_id=existing.id, uploaded_by_id=user.id, **att_meta,
                        ))
                dupes.append({"number": num, "project": project_code, "action": dup_action})
                continue

            if commit:
                inv = Invoice(
                    number=num, project_id=project.id, type=itype,
                    invoice_date=inv_date, due_date=due,
                    vendor_client_id=vc_id,
                    party_name=_str(first.get("party_name")),
                    subtotal=subtotal, tax=tax, total=total,
                    status=InvoiceStatus.DRAFT,
                    notes=_str(first.get("notes")),
                    created_by_id=user.id,
                )
                for d in items_data:
                    inv.items.append(InvoiceItem(**d))
                db.add(inv)
                if att_meta:
                    await db.flush()
                    db.add(InvoiceAttachment(
                        invoice_id=inv.id, uploaded_by_id=user.id, **att_meta,
                    ))
            new.append({"number": num, "items": len(items_data), "total": str(total)})
        except Exception as e:
            errors.append({"row": first_row_num, "message": str(e),
                           "raw": {"number": num}})
    return new, dupes, errors


async def import_purchase_orders(rows, db, user, *, commit, dup_action):
    """Grouped by `_po_ref` (kunci di sheet, BUKAN nomor PO).
    Kalau kolom `number` diisi, dipakai sebagai natural key (project_id, number);
    kalau kosong, PO selalu dianggap baru dan nomor otomatis dibuatkan."""
    new, dupes, errors = [], [], []
    groups: dict[str, list[tuple[int, dict]]] = {}
    for i, r in enumerate(rows, start=2):
        ref = _str(r.get("_po_ref"))
        if not ref:
            errors.append({"row": i, "message": "_po_ref wajib (untuk grouping item PO yang sama)", "raw": r})
            continue
        groups.setdefault(ref, []).append((i, r))

    for ref, group in groups.items():
        first_row_num, first = group[0]
        try:
            project_code = _str(first.get("project_code"))
            project = await _lookup(db, Project, code=project_code)
            if not project:
                raise ValueError(f"Proyek '{project_code}' tidak ditemukan")
            company_name = _str(first.get("company_name"))
            company = await _lookup_ci(db, Company, "name", company_name) if company_name else None
            if not company:
                raise ValueError(f"Perusahaan '{company_name}' tidak ditemukan")
            po_date = _parse_date(first.get("po_date"))
            if not po_date:
                raise ValueError("po_date wajib")
            needed = _parse_date(first.get("needed_date"))
            tax = _parse_decimal(first.get("tax") or 0)
            discount = _parse_decimal(first.get("discount") or 0)
            vc_id = None
            if vn := _str(first.get("vendor_client_name")):
                vc = await _lookup_ci(db, VendorClient, "name", vn)
                if vc:
                    vc_id = vc.id

            items_data: list[dict] = []
            for row_no, r in group:
                desc = _str(r.get("item_description"))
                if not desc:
                    continue
                qty = _parse_decimal(r.get("item_quantity") or 1)
                price = _parse_decimal(r.get("item_unit_price") or 0)
                items_data.append(dict(
                    description=desc, quantity=qty,
                    unit=_str(r.get("item_unit")),
                    unit_price=price, subtotal=price * qty,
                ))
            subtotal = sum((it["subtotal"] for it in items_data), Decimal("0"))
            total = subtotal + tax - discount

            explicit_number = _str(first.get("number"))
            existing = None
            if explicit_number:
                existing = (await db.execute(
                    select(PurchaseOrder)
                    .options(selectinload(PurchaseOrder.items))
                    .where(
                        PurchaseOrder.project_id == project.id,
                        PurchaseOrder.number == explicit_number,
                        PurchaseOrder.deleted_at.is_(None),
                    )
                )).scalar_one_or_none()

            if existing:
                if dup_action == "error":
                    raise ValueError(f"PO '{explicit_number}' di proyek {project_code} sudah ada")
                if dup_action == "update" and commit:
                    existing.vendor_client_id = vc_id or existing.vendor_client_id
                    existing.vendor_name = _str(first.get("vendor_name")) or existing.vendor_name
                    existing.po_date = po_date
                    existing.needed_date = needed
                    existing.tax = tax
                    existing.discount = discount
                    existing.subtotal = subtotal
                    existing.total = total
                    existing.payment_terms = _str(first.get("payment_terms")) or existing.payment_terms
                    existing.notes = _str(first.get("notes")) or existing.notes
                    existing.items.clear()
                    await db.flush()
                    for d in items_data:
                        existing.items.append(POItem(**d))
                dupes.append({"number": explicit_number, "_po_ref": ref, "action": dup_action})
                continue

            if commit:
                # generate atau pakai eksplisit
                if explicit_number:
                    number = explicit_number
                else:
                    prefix = f"PO/{po_date.year}/{po_date.month:02d}/{project.code.upper()}/"
                    count_q = (await db.execute(
                        select(func.count()).select_from(PurchaseOrder).where(
                            PurchaseOrder.company_id == company.id,
                            PurchaseOrder.number.like(f"{prefix}%"),
                        )
                    )).scalar_one()
                    number = f"{prefix}{count_q + 1:04d}"
                po = PurchaseOrder(
                    number=number, project_id=project.id, company_id=company.id,
                    vendor_client_id=vc_id, vendor_name=_str(first.get("vendor_name")),
                    po_date=po_date, needed_date=needed,
                    subtotal=subtotal, tax=tax, discount=discount, total=total,
                    payment_terms=_str(first.get("payment_terms")),
                    notes=_str(first.get("notes")),
                    status=POStatus.DRAFT,
                    created_by_id=user.id,
                )
                for d in items_data:
                    po.items.append(POItem(**d))
                db.add(po)
            new.append({"_po_ref": ref, "items": len(items_data), "total": str(total)})
        except Exception as e:
            errors.append({"row": first_row_num, "message": str(e), "raw": {"_po_ref": ref}})
    return new, dupes, errors


# ---------- Registry ----------
# Signature: handler(rows, db, user, *, commit: bool, dup_action: str)
#   -> (new: list, dupes: list, errors: list)
EntityHandler = Callable[..., Awaitable[tuple[list, list, list]]]


SCHEMAS: dict[str, dict[str, Any]] = {
    "companies": {
        "label": "Perusahaan",
        "headers": ["name", "address", "npwp", "phone", "email", "director_name", "bank_account"],
        "example": ["PT Contoh Sejahtera", "Jl. Demo No. 1", "01.234.567.8-091.000",
                    "021-1234567", "info@contoh.co.id", "Direktur Utama",
                    "BCA 123-456-789"],
        "handler": import_companies,
    },
    "categories": {
        "label": "Kategori",
        "headers": ["name", "type", "description"],
        "example": ["Material Bangunan", "OUT", "Semen, besi, dll"],
        "handler": import_categories,
    },
    "vendors-clients": {
        "label": "Vendor / Client",
        "headers": ["name", "type", "contact", "phone", "email", "npwp", "address", "bank_account"],
        "example": ["Toko Bangunan Sentosa", "VENDOR", "Pak Joko", "021-555-1111",
                    "joko@sentosa.co.id", "", "", "BCA 222-333-444"],
        "handler": import_vendors_clients,
    },
    "projects": {
        "label": "Proyek",
        "headers": ["code", "name", "location", "company_name", "pic_email",
                    "start_date", "end_date", "status",
                    "project_value", "budget_amount",
                    "currency", "overbudget_tolerance_pct", "notes",
                    "document_url", "document_label"],
        "example": ["PRJ-DEMO", "Proyek Demo", "Jakarta",
                    "PT Contoh Sejahtera", "budi@bintang.me",
                    "2026-04-01", "2026-12-31", "AKTIF",
                    357000000, 250000000,
                    "IDR", 5, "Catatan demo",
                    "https://drive.google.com/file/d/abc123/view", "Kontrak"],
        "note": (
            "project_value = nilai kontrak/SPK. budget_amount = target pengeluaran. "
            "Kalau budget_amount dikosongkan, otomatis dihitung 70% dari project_value. "
            "document_url opsional -- isi URL dokumen (mis. Google Drive yg sudah di-share) "
            "untuk auto-attach sebagai dokumen proyek; document_label = 'Kontrak', 'BAST', dll."
        ),
        "handler": import_projects,
    },
    "transactions": {
        "label": "Transaksi",
        "headers": ["project_code", "tx_date", "type", "category_name", "amount",
                    "party_name", "party_type", "vendor_client_name",
                    "payment_method", "reference_no", "description",
                    "attachment_url", "attachment_label"],
        "example": ["PRJ-001", "2026-04-15", "OUT", "Material Bangunan", 5500000,
                    "Toko Bangunan Sentosa", "COMPANY", "Toko Bangunan Sentosa",
                    "TRANSFER", "TRF-2026-001", "Pembelian semen 50 sak",
                    "https://drive.google.com/file/d/xyz789/view", "Bukti TF BCA"],
        "note": (
            "attachment_url opsional -- isi URL bukti (mis. Google Drive yg sudah di-share) "
            "untuk auto-attach sebagai bukti transaksi. attachment_label = nama dokumen "
            "(opsional). Kompatibel juga dengan kolom 'bukti_url' dan 'bukti_label'."
        ),
        "handler": import_transactions,
    },
    "invoices": {
        "label": "Invoice",
        "headers": ["number", "project_code", "type", "invoice_date", "due_date",
                    "vendor_client_name", "party_name", "tax", "notes",
                    "item_description", "item_quantity", "item_unit", "item_unit_price",
                    "attachment_url", "attachment_label"],
        "example": ["INV/2026/04/PRJ-001/0099", "PRJ-001", "OUT", "2026-04-20",
                    "2026-05-20", "PT Klien Sukses Makmur", "PT Klien Sukses Makmur",
                    11000000, "Termin demo", "Pekerjaan struktur", 1, "lot", 100000000,
                    "https://drive.google.com/file/d/inv001/view", "Scan invoice"],
        "handler": import_invoices,
        "note": (
            "Untuk invoice dengan banyak item: ulang baris dengan number yang sama, isi "
            "kolom item_* berbeda. attachment_url & attachment_label dibaca dari baris "
            "PERTAMA grup -- isi cukup di baris pertama saja."
        ),
    },
    "purchase-orders": {
        "label": "Purchase Order",
        "headers": ["_po_ref", "project_code", "company_name", "vendor_client_name",
                    "vendor_name", "po_date", "needed_date", "tax", "discount",
                    "payment_terms", "notes",
                    "item_description", "item_quantity", "item_unit", "item_unit_price"],
        "example": ["PO-1", "PRJ-001", "PT Bintang Karya Abadi",
                    "CV Mitra Beton Pratama", "CV Mitra Beton Pratama",
                    "2026-04-15", "2026-04-25", 0, 0, "NET 30", "",
                    "Beton K-300 ready mix", 50, "m3", 400000],
        "handler": import_purchase_orders,
        "note": "Pakai _po_ref unik per PO. Tambah baris dengan _po_ref sama untuk item lain.",
    },
}
