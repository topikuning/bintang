import re
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# Pola: angka Indonesia formatted (mis. "1.234.567" atau "1.234.567,89")
# ATAU angka biasa "1234567" / "1234567.89". Bisa ada tanda minus '-' atau '−'.
# Tolerant terhadap NBSP/spasi normal -- kadang ada di output _fmt_idr.
# Order alternasi penting: multi-thousand pattern (id-ID dot, en-US comma)
# DULU, baru fallback ke single-separator. Karena \d+(?:,\d+)? greedy bisa
# "menelan" sebagian dr 1,234,567.89 dan gagal anchor $.
_NUM_RE = re.compile(
    r"""
    ^\s*                                # leading spaces
    [−\-]?                              # optional minus (- atau en-dash)
    \s*
    (
        \d{1,3}(?:\.\d{3})+(?:,\d+)?    # 1.234.567 atau 1.234.567,89 (id-ID)
        |
        \d{1,3}(?:,\d{3})+(?:\.\d+)?    # 1,234,567 atau 1,234,567.89 (en-US)
        |
        \d+(?:,\d+)?                    # 1234 atau 1234,89
        |
        \d+(?:\.\d+)?                   # 1234 atau 1234.89
    )
    \s*$
    """,
    re.VERBOSE,
)


def _try_parse_number(v: Any) -> float | int | None:
    """Konversi nilai cell ke angka jika tampak seperti formatted number.

    Kalau bukan -- mis. teks biasa, tanggal, label "TOTAL" -- return None
    (cell tetap tertulis as-is). Hanya angka pure (tanpa prefix 'Rp')
    yang berhasil di-parse, karena _fmt_idr biasanya wrap di kolom totals
    dgn 'Rp' di label, sedangkan di cell angka di body table sudah polos.
    """
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return v
    if not isinstance(v, str):
        return None
    s = v.strip()
    if not s:
        return None
    if not _NUM_RE.match(s):
        return None
    # Normalize minus
    s = s.replace("−", "-").replace(" ", " ").replace(" ", "")
    # Deteksi locale: kalau ada koma SETELAH titik, anggap en-US.
    # Kalau ada titik DAN koma, koma = desimal (id-ID).
    # Kalau hanya satu separator: tergantung posisi.
    has_dot = "." in s
    has_comma = "," in s
    try:
        if has_dot and has_comma:
            # id-ID: titik = ribuan, koma = desimal
            return float(s.replace(".", "").replace(",", "."))
        if has_comma and not has_dot:
            # bisa "1,5" (desimal id) atau "1,234" (ribuan en-US). Heuristik:
            # kalau 3 digit di belakang koma, anggap ribuan en-US. Else desimal.
            after = s.rsplit(",", 1)[1]
            if len(after) == 3:
                return float(s.replace(",", ""))
            return float(s.replace(",", "."))
        if has_dot and not has_comma:
            # bisa "1.5" (desimal) atau "1.234" (ribuan id). Heuristik sama.
            after = s.rsplit(".", 1)[1]
            if len(after) == 3:
                return float(s.replace(".", ""))
            return float(s)
        return float(s)
    except (ValueError, IndexError):
        return None


def build_xlsx(
    title: str,
    headers: list[str],
    rows: list[list[Any]],
    *,
    subtitle: str | None = None,
    filters: dict | None = None,
    totals: dict | None = None,
    cols: list[dict] | None = None,
    sheet_name: str = "Report",
) -> bytes:
    """Build Excel report. Cell yg di-detect sbg angka (numeric format
    Indonesia atau en-US) di-write sbg raw number + Excel number format
    `#,##0` supaya:
      1. User bisa SUM / formula di Excel
      2. Lokalisasi separator mengikuti setting Excel client (id-ID /
         en-US tetap tampil benar)
      3. Tidak ada warning 'angka disimpan sbg teks'
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E5E7EB")
    # Format angka: thousand separator + dec 2 jika perlu (#,##0.##). Excel
    # otomatis tampil dgn separator locale user.
    num_format = "#,##0.##"
    # Kolom yg align='num' diharapkan angka. Tandai utk apply format.
    num_col_idx = {i for i, c in enumerate(cols or []) if c.get("align") == "num"}

    ws.append([title])
    ws[ws.max_row][0].font = Font(bold=True, size=14)
    if subtitle:
        ws.append([subtitle])
    if filters:
        ws.append([
            "Filter: "
            + " | ".join(f"{k}: {v}" for k, v in filters.items())
        ])
    ws.append([])
    header_row_idx = ws.max_row + 1
    ws.append(headers)
    for cell in ws[header_row_idx]:
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        # Write row cell-by-cell supaya bisa apply number format ke numeric cell.
        ws_row_idx = ws.max_row + 1
        ws.append([""] * len(row))  # placeholder
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=ws_row_idx, column=col_idx)
            # Coba parse ke angka kalau kolomnya numeric ATAU value tampak
            # spt angka. Conservative: utk kolom num, force coba; utk kolom
            # lain, hanya jika RE match (mengurangi false positive).
            is_num_col = (col_idx - 1) in num_col_idx
            parsed = _try_parse_number(val) if is_num_col or isinstance(val, str) else None
            if parsed is not None:
                cell.value = parsed
                cell.number_format = num_format
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.value = val if val is not None else ""

    if totals:
        ws.append([])
        for k, v in totals.items():
            ws.append([k, v])
            ws[ws.max_row][0].font = bold
            # Coba angka utk kolom kedua di total row.
            parsed = _try_parse_number(v)
            if parsed is not None:
                ws.cell(row=ws.max_row, column=2).value = parsed
                ws.cell(row=ws.max_row, column=2).number_format = num_format

    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(headers[col_idx - 1]))
        for r in rows:
            try:
                max_len = max(max_len, len(str(r[col_idx - 1])))
            except IndexError:
                pass
        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 50)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
